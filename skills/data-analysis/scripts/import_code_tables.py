"""공공데이터포털 참고문서(엑셀)의 코드표를 `reference/` 로 옮긴다.

**왜 이 스크립트가 필요한가** — `originTrialHall/dealings` 는 `trhl_cd`(공판장 코드)
없이는 어떤 조건으로도 0건을 반환한다. 그리고 그 코드는 **10자리**여서
(`7428200508` 원주원예농협공판장) 추측으로 찾을 수 없다. 코드표는 데이터 상세
페이지의 참고문서 엑셀 안에만 있다.

    출처: 공공데이터포털 15156054 「전국 산지공판장 거래정보」 참고문서
          (참고)전국 산지공판장 거래정보_코드.xlsx  · 이용허락범위 제한 없음

엑셀 원본은 저장소에 두지 않는다. 이 스크립트로 CSV만 뽑아 `reference/` 에 커밋한다.

    python scripts/import_code_tables.py <엑셀경로>
"""
from __future__ import annotations

import csv
import sys
import pathlib

sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent))
from api import use_utf8_stdout  # noqa: E402

import paths                                                        # noqa: E402

ROOT = paths.SKILL
REF = paths.REFERENCE

# 시트명 → (출력 파일명, 한국어 헤더)
#
# 두 참고문서를 함께 다룬다. 산지·도매는 `gds_*` 계열, 소매는 `perDay` 계열로
# 코드 체계가 갈리므로(DATA_SOURCES §1) 양쪽 코드표가 모두 있어야 3단계가 이어진다.
SHEETS = {
    # 15156054 「전국 산지공판장 거래정보」 — 산지·도매 공통 `gds_*`
    "공판장코드": ("trial_halls.csv", ["공판장코드", "공판장명"]),
    "상품대분류코드": ("goods_large.csv", ["대분류코드", "대분류명"]),
    "상품중분류코드": ("goods_medium.csv", ["대분류코드", "중분류코드", "중분류명"]),
    "상품소분류코드": ("goods_small.csv", ["대분류코드", "중분류코드", "소분류코드", "소분류명"]),

    # 15156057 「일별 도·소매 가격정보」 — 소매 `perDay`
    "부류코드": ("perday_category.csv", ["부류코드", "부류명"]),
    "품목코드": ("perday_items.csv", ["부류코드", "품목코드", "품목명"]),
    "품종코드": ("perday_varieties.csv", ["부류코드", "품목코드", "품종코드", "품종명"]),
    "구분코드": ("perday_division.csv", ["구분코드", "구분명"]),
    "시장코드": ("perday_markets.csv", ["시군구코드", "시장코드", "시장명"]),
}


def main() -> None:
    use_utf8_stdout()
    if len(sys.argv) < 2:
        raise SystemExit(
            "사용법: python scripts/import_code_tables.py <코드표.xlsx> [코드표2.xlsx ...]")

    try:
        import openpyxl
    except ImportError:
        raise SystemExit("openpyxl 이 필요합니다: pip install openpyxl")

    REF.mkdir(parents=True, exist_ok=True)
    for arg in sys.argv[1:]:
        src = pathlib.Path(arg)
        if not src.exists():
            print(f"  ⚠️ 파일 없음: {src}")
            continue
        print(f"[{src.name}]")
        옮기기(openpyxl.load_workbook(src, read_only=True))

    print("\n  출처: 공공데이터포털 15156054 · 15156057 참고문서 · 이용허락범위 제한 없음")


def 옮기기(workbook) -> None:
    for sheet_name, (out_name, header) in SHEETS.items():
        if sheet_name not in workbook.sheetnames:
            continue

        sheet = workbook[sheet_name]
        rows = []
        for index, values in enumerate(sheet.iter_rows(values_only=True)):
            if index == 0:                       # 엑셀 헤더는 버리고 우리 헤더를 쓴다
                continue
            if values is None or values[0] is None:
                continue
            # 코드는 앞자리 0이 의미를 가지므로 반드시 문자열로 다룬다
            rows.append([("" if v is None else str(v).strip()) for v in values[:len(header)]])

        out = REF / out_name
        with out.open("w", encoding="utf-8-sig", newline="") as fp:
            writer = csv.writer(fp)
            writer.writerow(header)
            writer.writerows(rows)
        print(f"  → reference/{out_name}  {len(rows)}행")


if __name__ == "__main__":
    main()
